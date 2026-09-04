#!/usr/bin/env python3
"""Prepare the Maxis Personal Key UAT event without emitting human keys.

This generator is deliberately execution-free.  It validates the authoritative
workbook and writes transaction-safe setup and rollback-only certification SQL.
Neither SQL file contains a raw Personal Key or an opaque derived credential;
the setup contains only the frozen Team Formation enrollment hashes.
"""
from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
import hashlib
import json
from pathlib import Path
import re
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from services.personal_key_credentials import (
    derive_personal_key_credential,
    normalize_personal_key,
    team_formation_credential_hash,
)


EVENT_ID = "MAXIS-UAT-PREASSIGNED"
JOIN_CODE = "MXKEY7"
EVENT_NAME = "Maxis Mission AI — Personal Key UAT"
ACTOR = "maxis_personal_key_uat_setup"
EXPECTED_WORKBOOK_SHA256 = "ff9f2070d4c5a94b7666828bcd4731ea0a385117d6512937a11cf02ce3819640"
EXPECTED_COUNTS = {1: 11, 2: 11, 3: 11, 4: 11, 5: 12, 6: 12}
TEAM_CONFIG = {
    1: ("Japan", "🇯🇵", "KONNICHIWA!"),
    2: ("South Korea", "🇰🇷", "ANNYEONGHASEYO!"),
    3: ("France", "🇫🇷", "BONJOUR!"),
    4: ("Italy", "🇮🇹", "CIAO!"),
    5: ("Brazil", "🇧🇷", "OLÁ!"),
    6: ("Thailand", "🇹🇭", "SAWASDEE!"),
}
TEAM_PATTERN = re.compile(r"^Team ([1-6])$")


@dataclass(frozen=True)
class RosterMember:
    name: str
    team_number: int
    personal_key: str

    @property
    def team_id(self) -> str:
        return f"{EVENT_ID}-TEAM-{self.team_number:02d}"

    @property
    def country(self) -> str:
        return TEAM_CONFIG[self.team_number][0]

    @property
    def derived_credential(self) -> str:
        return derive_personal_key_credential(EVENT_ID, self.personal_key)

    @property
    def enrollment_credential_hash(self) -> str:
        return team_formation_credential_hash(self.derived_credential)


def _sheet_rows(workbook, sheet_name: str) -> list[RosterMember]:
    sheet = workbook[sheet_name]
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))]
    required = ("Participant Name", "Team", "Personal Key")
    if any(header not in headers for header in required):
        raise ValueError(f"{sheet_name} is missing required roster headers.")
    name_index = headers.index("Participant Name")
    team_index = headers.index("Team")
    key_index = headers.index("Personal Key")
    members: list[RosterMember] = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if len(row) <= max(name_index, team_index, key_index) or row[name_index] is None:
            continue
        name = str(row[name_index])
        team_label = str(row[team_index])
        raw_key = str(row[key_index])
        match = TEAM_PATTERN.fullmatch(team_label)
        if not match:
            raise ValueError(f"Invalid authoritative team label for roster row {len(members) + 5}.")
        if name != name.strip() or not name:
            raise ValueError(f"Participant name has invalid surrounding whitespace at row {len(members) + 5}.")
        normalized_key = normalize_personal_key(raw_key)
        if normalized_key != raw_key:
            raise ValueError(f"Personal Key is not already authoritative uppercase form at row {len(members) + 5}.")
        members.append(RosterMember(name=name, team_number=int(match.group(1)), personal_key=raw_key))
    return members


def load_authoritative_roster(workbook_path: Path) -> list[RosterMember]:
    source_digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    if source_digest != EXPECTED_WORKBOOK_SHA256:
        raise ValueError("Workbook SHA-256 does not match the approved authoritative source.")
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    required_sheets = {"Facilitator Master", "A-Z Lookup"}
    if not required_sheets.issubset(workbook.sheetnames):
        raise ValueError("Workbook is missing an authoritative roster sheet.")
    roster = _sheet_rows(workbook, "Facilitator Master")
    lookup = _sheet_rows(workbook, "A-Z Lookup")
    if len(roster) != 68 or Counter(member.team_number for member in roster) != Counter(EXPECTED_COUNTS):
        raise ValueError("Workbook does not reconcile to the approved 68-person country counts.")
    if len({member.name for member in roster}) != 68:
        raise ValueError("Participant names must be unique.")
    if len({member.personal_key for member in roster}) != 68:
        raise ValueError("Personal Keys must be unique.")
    if {(m.name, m.team_number, m.personal_key) for m in roster} != {
        (m.name, m.team_number, m.personal_key) for m in lookup
    }:
        raise ValueError("A-Z Lookup does not reconcile with Facilitator Master.")
    if len({member.derived_credential for member in roster}) != 68:
        raise ValueError("Derived credentials must be unique within the event.")
    if len({member.enrollment_credential_hash for member in roster}) != 68:
        raise ValueError("Enrollment credential hashes must be unique within the event.")
    return roster


def _sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _json_literal(value) -> str:
    return _sql_literal(json.dumps(value, ensure_ascii=False, separators=(",", ":"))) + "::jsonb"


def _team_rows() -> list[dict]:
    return [
        {
            "team_id": f"{EVENT_ID}-TEAM-{team_number:02d}",
            "team_name": country,
            "country": country,
            "team_flag": flag,
        }
        for team_number, (country, flag, _greeting) in TEAM_CONFIG.items()
    ]


def _event_metadata() -> dict:
    return {
        "Client": "Maxis",
        "NumberOfTeams": 6,
        "PersonalKeyExperience": {
            "SchemaVersion": 1,
            "CredentialDerivation": "EXOS_TEAM_FORMATION_PERSONAL_KEY_V1",
        },
        "TeamIdentityConfig": {
            "ThemeType": "COUNTRY",
            "ThemeName": "Countries",
            "Identities": [
                {
                    "TeamID": f"{EVENT_ID}-TEAM-{team_number:02d}",
                    "TeamIdentity": country,
                    "Country": country,
                    "Emoji": flag,
                    "Greeting": greeting,
                }
                for team_number, (country, flag, greeting) in TEAM_CONFIG.items()
            ],
        },
    }


def build_setup_sql(roster: list[RosterMember]) -> str:
    roster_payload = [
        {
            "EnrollmentCredentialHash": member.enrollment_credential_hash,
            "DisplayName": member.name,
            "TeamID": member.team_id,
        }
        for member in roster
    ]
    capacities = {
        f"{EVENT_ID}-TEAM-{team_number:02d}": capacity
        for team_number, capacity in EXPECTED_COUNTS.items()
    }
    metadata = _event_metadata()
    team_counts = " union all ".join(
        f"select {_sql_literal(f'{EVENT_ID}-TEAM-{team_number:02d}')}::text as team_id, {count}::bigint as expected_count"
        for team_number, count in EXPECTED_COUNTS.items()
    )
    sql = f"""-- Generated Maxis Personal Key UAT setup: contains hashes, never human or derived credentials.
begin;

do $guard$
begin
    if exists (select 1 from public.events_v2 where event_id = {_sql_literal(EVENT_ID)}) then
        raise exception 'Target UAT event already exists';
    end if;
    if exists (select 1 from public.events_v2 where join_code = {_sql_literal(JOIN_CODE)}) then
        raise exception 'Target UAT join code already exists';
    end if;
end;
$guard$;

select public.exos_v2_publish_event(
    {_sql_literal(EVENT_ID)},
    {_sql_literal(JOIN_CODE)},
    {_sql_literal(EVENT_NAME)},
    {_json_literal(_team_rows())},
    'TEAM_COMPETITIVE'::public.exos_v2_scoring_mode,
    'STANDARD'
);

update public.events_v2
   set event_payload = coalesce(event_payload, '{{}}'::jsonb) || {_json_literal(metadata)},
       lifecycle_status = 'PUBLISHED',
       updated_at = now()
 where event_id = {_sql_literal(EVENT_ID)};

select public.exos_v2_configure_team_formation(
    {_sql_literal(EVENT_ID)},
    'PREASSIGNED',
    {_json_literal(capacities)},
    {_json_literal(roster_payload)},
    {_sql_literal(ACTOR)}
);

select public.exos_v2_open_team_formation({_sql_literal(EVENT_ID)}, {_sql_literal(ACTOR)});

do $assert$
declare
    v_bad_count bigint;
begin
    if (select count(*) from public.events_v2 where event_id = {_sql_literal(EVENT_ID)} and join_code = {_sql_literal(JOIN_CODE)}) <> 1 then
        raise exception 'Event identity assertion failed';
    end if;
    if (select event_payload #>> '{{TeamFormation,Mode}}' from public.events_v2 where event_id = {_sql_literal(EVENT_ID)}) <> 'PREASSIGNED'
       or (select event_payload #>> '{{TeamFormation,Phase}}' from public.events_v2 where event_id = {_sql_literal(EVENT_ID)}) <> 'REGISTRATION_OPEN' then
        raise exception 'Registration-open assertion failed';
    end if;
    if (select count(*) from public.participants_v2 where event_id = {_sql_literal(EVENT_ID)} and not is_archived) <> 68 then
        raise exception 'Roster count assertion failed';
    end if;
    if (select count(distinct enrollment_credential_hash) from public.participants_v2 where event_id = {_sql_literal(EVENT_ID)} and not is_archived) <> 68 then
        raise exception 'Enrollment hash uniqueness assertion failed';
    end if;
    if exists (
        select 1 from public.participants_v2
         where event_id = {_sql_literal(EVENT_ID)}
           and (participant_status <> 'PREASSIGNED'
                or enrollment_credential_hash !~ '^[0-9a-f]{{64}}$'
                or participant_payload::text ~* 'personal.?key|derived.?credential')
    ) then
        raise exception 'Roster privacy/status assertion failed';
    end if;
    select count(*) into v_bad_count
      from ({team_counts}) expected
      left join lateral (
          select count(*)::bigint as actual_count
            from public.participants_v2 p
           where p.event_id = {_sql_literal(EVENT_ID)}
             and p.team_id = expected.team_id
             and not p.is_archived
      ) actual on true
     where actual.actual_count <> expected.expected_count;
    if v_bad_count <> 0 then
        raise exception 'Country count assertion failed';
    end if;
end;
$assert$;

commit;
"""
    forbidden = [member.personal_key for member in roster] + [member.derived_credential for member in roster]
    if any(secret in sql for secret in forbidden):
        raise RuntimeError("Generated SQL contains a raw or derived credential.")
    return sql


def build_certification_sql(roster: list[RosterMember]) -> str:
    representatives = [next(member for member in roster if member.team_number == team) for team in TEAM_CONFIG]
    wrong_key = next(
        candidate
        for candidate in (f"ZZ{number:04d}" for number in range(10_000))
        if candidate not in {member.personal_key for member in roster}
    )
    claims = []
    for member in representatives:
        claims.append(
            f"""
    v_identity := public.exos_v2_team_formation_claim_preassigned(
        {_sql_literal(JOIN_CODE)}, {_sql_literal(member.derived_credential)},
        {_sql_literal(f'MAXIS-CERT-DEVICE-{member.team_number}')}
    );
    if v_identity->>'EventID' <> {_sql_literal(EVENT_ID)}
       or v_identity->>'Name' <> {_sql_literal(member.name)}
       or v_identity->>'Country' <> {_sql_literal(member.country)}
       or v_identity->>'TeamID' <> {_sql_literal(member.team_id)} then
        raise exception 'Valid-key country assertion failed for country {member.team_number}';
    end if;
"""
        )
    first = representatives[0]
    sql = f"""-- Generated Maxis Personal Key certification. All mutations roll back.
begin;

do $cert$
declare
    v_identity jsonb;
    v_repeat jsonb;
    v_first_participant_id text;
    v_first_session_token text;
    v_before_participants bigint;
    v_before_sessions bigint;
    v_before_team_rows bigint;
begin
    select count(*) into v_before_participants from public.participants_v2 where event_id = {_sql_literal(EVENT_ID)};
    select count(*) into v_before_sessions from public.participant_sessions_v2 where event_id = {_sql_literal(EVENT_ID)};
    select count(*) into v_before_team_rows from public.teams_v2 where event_id = {_sql_literal(EVENT_ID)};
    {claims[0]}
    v_first_participant_id := v_identity->>'ParticipantID';
    v_first_session_token := v_identity->>'SessionToken';
    {''.join(claims[1:])}
    v_repeat := public.exos_v2_team_formation_claim_preassigned(
        {_sql_literal(JOIN_CODE)}, {_sql_literal(first.derived_credential)}, 'MAXIS-CERT-DEVICE-1'
    );
    if v_repeat->>'ParticipantID' <> v_first_participant_id
       or v_repeat->>'SessionToken' <> v_first_session_token
       or v_repeat->>'Name' <> {_sql_literal(first.name)} then
        raise exception 'Same-device identity assertion failed';
    end if;
    if coalesce((v_repeat->>'Idempotent')::boolean, false) is not true then
        raise exception 'Same-device idempotency assertion failed';
    end if;
    if (select count(*) from public.participants_v2 where event_id = {_sql_literal(EVENT_ID)}) <> v_before_participants then
        raise exception 'Duplicate participant assertion failed';
    end if;
    begin
        perform public.exos_v2_team_formation_claim_preassigned(
            {_sql_literal(JOIN_CODE)}, {_sql_literal(derive_personal_key_credential(EVENT_ID, wrong_key))}, 'MAXIS-CERT-WRONG-KEY'
        );
        raise exception 'Wrong key unexpectedly authenticated';
    exception
        when others then
            if sqlerrm <> 'PREASSIGNED_ENROLLMENT_NOT_FOUND' then
                raise;
            end if;
    end;
    if (select count(*) from public.participant_sessions_v2 where event_id = {_sql_literal(EVENT_ID)}) <> v_before_sessions + 6 then
        raise exception 'Session/idempotency count assertion failed';
    end if;
    if (select count(*) from public.teams_v2 where event_id = {_sql_literal(EVENT_ID)}) <> v_before_team_rows then
        raise exception 'Team mutation assertion failed';
    end if;
end;
$cert$;

create temporary table maxis_personal_key_anon_result(payload jsonb) on commit drop;
grant insert, select on table maxis_personal_key_anon_result to anon;
set local role anon;
insert into maxis_personal_key_anon_result(payload)
select public.exos_v2_team_formation_claim_preassigned(
    {_sql_literal(JOIN_CODE)}, {_sql_literal(first.derived_credential)}, 'MAXIS-CERT-DEVICE-1'
);
reset role;

do $anon_assert$
begin
    if not exists (
        select 1 from maxis_personal_key_anon_result
         where payload->>'EventID' = {_sql_literal(EVENT_ID)}
           and payload->>'Name' = {_sql_literal(first.name)}
           and payload->>'Country' = {_sql_literal(first.country)}
           and coalesce((payload->>'Idempotent')::boolean, false)
    ) then
        raise exception 'Anonymous RPC execution/identity assertion failed';
    end if;
end;
$anon_assert$;

rollback;
"""
    forbidden = [member.personal_key for member in roster]
    if any(secret in sql for secret in forbidden):
        raise RuntimeError("Generated certification SQL contains a raw Personal Key.")
    return sql


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--setup-sql", required=True, type=Path)
    parser.add_argument("--certification-sql", required=True, type=Path)
    args = parser.parse_args()
    roster = load_authoritative_roster(args.workbook)
    args.setup_sql.write_text(build_setup_sql(roster), encoding="utf-8")
    args.certification_sql.write_text(build_certification_sql(roster), encoding="utf-8")
    print("MAXIS_PERSONAL_KEY_PACKAGE_READY rows=68 countries=6 raw_keys_emitted=0 derived_credentials_emitted=0")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
